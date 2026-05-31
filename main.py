import customtkinter as ctk
import sqlite3
from pathlib import Path
from tkinter import messagebox
from PIL import Image
import sys
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import locale
from queue import Queue
from threading import Lock
import time

try:
    locale.setlocale(locale.LC_TIME, "pt_BR.UTF-8")
except Exception:
    try:
        locale.setlocale(locale.LC_TIME, "Portuguese_Brazil.1252")
    except Exception:
        pass

BASE_DIR   = Path(__file__).parent.absolute()
DB_PATH    = BASE_DIR / "biblioteca.db"
LOGO_PATH  = BASE_DIR / "erasebg-transformed.png"
EXPORT_DIR = BASE_DIR / "Relatorios"
EXPORT_DIR.mkdir(exist_ok=True)

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

TURMAS_OPCOES   = ["1", "2", "3"]
ESTANTES_LETRAS = ["A", "B", "C", "D", "E", "F", "G", "H"]
FILEIRAS_NUM    = ["1", "2", "3", "4", "5"]

MESES_PT = {
    1:"JANEIRO", 2:"FEVEREIRO", 3:"MARÇO",    4:"ABRIL",
    5:"MAIO",    6:"JUNHO",     7:"JULHO",     8:"AGOSTO",
    9:"SETEMBRO",10:"OUTUBRO", 11:"NOVEMBRO", 12:"DEZEMBRO"
}

C = {
    "amarelo":      "#FFCC00",
    "vermelho":     "#A31C1C",
    "vermelho_esc": "#7A1515",
    "verde":        "#2E7D32",
    "verde_esc":    "#1B5E20",
    "fundo":        "#F4F4F6",
    "card":         "#FFFFFF",
    "borda":        "#E8E8EC",
    "texto":        "#1A1A2E",
    "subtexto":     "#6B7280",
    "cinza_btn":    "#6B7280",
    "cinza_hover":  "#4B5563",
    "perigo":       "#B71C1C",
    "perigo_hover": "#7F0000",
    "inativo":      "#E5E7EB",
    "inativo_txt":  "#374151",
}

FONTE_TITULO  = ("Segoe UI", 20, "bold")
FONTE_SECAO   = ("Segoe UI", 14, "bold")
FONTE_CORPO   = ("Segoe UI", 13)
FONTE_PEQUENA = ("Segoe UI", 11)
FONTE_BOLD    = ("Segoe UI", 13, "bold")

def estilo_entry(width=180, **kw):
    return dict(width=width, height=38, corner_radius=10,
                border_color=C["borda"], border_width=1,
                fg_color="white", text_color=C["texto"],
                font=FONTE_CORPO, **kw)

def estilo_segmented(width=200):
    return dict(width=width, height=34,
                selected_color=C["vermelho"],
                selected_hover_color=C["vermelho_esc"],
                unselected_color=C["inativo"],
                unselected_hover_color="#D1D5DB",
                text_color="white", font=FONTE_BOLD, corner_radius=8)

def estilo_btn_primario(width=140, height=38, **kw):
    return dict(width=width, height=height, corner_radius=10,
                fg_color=C["vermelho"], hover_color=C["vermelho_esc"],
                text_color="white", font=FONTE_BOLD, **kw)

def estilo_btn_verde(width=140, height=38, **kw):
    return dict(width=width, height=height, corner_radius=10,
                fg_color=C["verde"], hover_color=C["verde_esc"],
                text_color="white", font=FONTE_BOLD, **kw)

def estilo_btn_cinza(width=120, height=38, **kw):
    return dict(width=width, height=height, corner_radius=10,
                fg_color=C["cinza_btn"], hover_color=C["cinza_hover"],
                text_color="white", font=FONTE_CORPO, **kw)

def cabecalho_tela(container, titulo, voltar_cmd):
    topo = ctk.CTkFrame(container, fg_color=C["vermelho"], corner_radius=0, height=58)
    topo.pack(fill="x"); topo.pack_propagate(False)
    ctk.CTkButton(topo, text="← Voltar", width=110, height=36,
                  fg_color="transparent", hover_color=C["vermelho_esc"],
                  text_color="white", font=("Segoe UI", 13),
                  command=voltar_cmd).pack(side="left", padx=18, pady=11)
    ctk.CTkLabel(topo, text=titulo, font=("Segoe UI", 16, "bold"),
                 text_color="white").pack(side="left", padx=6)

def separador(parent, pady=(0, 0)):
    ctk.CTkFrame(parent, height=1, fg_color=C["borda"]).pack(fill="x", pady=pady)

def lbl_sub(parent, texto):
    ctk.CTkLabel(parent, text=texto, font=FONTE_PEQUENA,
                 text_color=C["subtexto"]).pack(side="left", padx=(8, 2))

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

# ─────────────────────────────────────────────
#  BANCO DE DADOS
# ─────────────────────────────────────────────
def _conectar():
    conn = sqlite3.connect(
        DB_PATH,
        timeout=30,
        check_same_thread=False,
        isolation_level=None
    )
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA temp_store=MEMORY")
    conn.execute("PRAGMA cache_size=-10000")
    conn.execute("PRAGMA busy_timeout=30000")
    return conn

def _colunas_acervo():
    """Retorna lista de nomes de colunas da tabela acervo."""
    with _conectar() as conn:
        c = conn.cursor()
        c.execute("PRAGMA table_info(acervo)")
        return [r[1] for r in c.fetchall()]

def iniciar_db():
    try:
        with _conectar() as conn:
            cursor = conn.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS acervo (titulo TEXT, autor TEXT, genero TEXT)")
            cursor.execute("CREATE TABLE IF NOT EXISTS alunos (nome TEXT, serie TEXT, turma TEXT, turno TEXT)")
            cursor.execute("CREATE TABLE IF NOT EXISTS emprestimos (aluno_nome TEXT, livro_titulo TEXT, data_entrega TEXT, serie TEXT, turma TEXT, turno TEXT)")
            cursor.execute("CREATE TABLE IF NOT EXISTS ranking (nome TEXT PRIMARY KEY, lidos INTEGER DEFAULT 0, serie TEXT, turma TEXT, turno TEXT)")
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS turmas (
                    id    INTEGER PRIMARY KEY AUTOINCREMENT,
                    serie TEXT NOT NULL,
                    turma TEXT NOT NULL,
                    turno TEXT NOT NULL,
                    UNIQUE(serie, turma, turno)
                )
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_alunos_nome   ON alunos (nome)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_acervo_titulo ON acervo (titulo)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_alunos_turma  ON alunos (serie, turma, turno)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_emp_turma     ON emprestimos (serie, turma, turno)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_ranking_turma ON ranking (serie, turma, turno)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_emp_nome ON emprestimos(aluno_nome)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_emp_livro ON emprestimos(livro_titulo)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_acervo_codigo ON acervo(codigo)")

            def garantir_col(tabela, coluna, tipo="TEXT"):
                cursor.execute(f"PRAGMA table_info({tabela})")
                if coluna not in [r[1] for r in cursor.fetchall()]:
                    cursor.execute(f"ALTER TABLE {tabela} ADD COLUMN {coluna} {tipo}")

            for col in ["turma", "turno"]:          garantir_col("alunos", col)
            for col in ["serie", "turma", "turno"]: garantir_col("emprestimos", col)
            for col in ["serie", "turma", "turno"]: garantir_col("ranking", col)

            # ── Acervo: garante as colunas novas sem remover 'edição' ──
            # O banco real pode ter 'edição' — mantemos ela e adicionamos
            # estante, fileira, codigo SE ainda não existirem
            for col in ["estante", "fileira", "codigo"]:
                garantir_col("acervo", col)

            cursor.execute("""
                INSERT OR IGNORE INTO turmas (serie, turma, turno)
                SELECT DISTINCT serie, turma, turno FROM alunos
                WHERE serie != '' AND turma != '' AND turno != ''
            """)
            conn.commit()
    except sqlite3.Error as e:
        messagebox.showerror("Erro crítico",
                             f"Não foi possível inicializar o banco de dados.\n\n{e}")
        sys.exit(1)

def listar_turmas_db():
    with _conectar() as conn:
        c = conn.cursor()
        c.execute("SELECT serie, turma, turno FROM turmas ORDER BY CAST(serie AS INTEGER), turma, turno")
        return c.fetchall()

def turmas_da_serie(serie):
    with _conectar() as conn:
        c = conn.cursor()
        c.execute("SELECT DISTINCT turma FROM alunos WHERE serie=? AND turma IS NOT NULL AND turma!='' ORDER BY turma",
                  (serie.strip(),))
        rows = [r[0] for r in c.fetchall()]
    return rows if rows else TURMAS_OPCOES

def em_thread(func, callback=None, erro_callback=None):
    def _run():
        try:
            with DB_LOCK:
                resultado = func()
            if callback:
                app_ref.after(0, lambda: callback(resultado))
        except Exception as e:
            if erro_callback:
                app_ref.after(0, lambda: erro_callback(e))
            else:
                app_ref.after(0, lambda: messagebox.showerror("Erro", str(e)))
    ativos = threading.active_count()
    if ativos < THREAD_LIMIT:
        threading.Thread(target=_run, daemon=True).start()

app_ref = None
DB_LOCK = Lock()
THREAD_LIMIT = 4
thread_queue = Queue()

# ─────────────────────────────────────────────
#  APLICAÇÃO
# ─────────────────────────────────────────────
class BibliotecaApp(ctk.CTk):
    def __init__(self):
        global app_ref
        super().__init__()
        app_ref = self
        self.title("Sistema Dom Mota - v8.5")
        self.resizable(True, True)
        self.after(100, lambda: self.focus_force())
        if sys.platform.startswith("linux"):
            self.attributes("-zoomed", True)
        else:
            self.state("zoomed")

        iniciar_db()
        self.configure(fg_color=C["fundo"])
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self.topo = ctk.CTkFrame(self, height=130, fg_color=C["amarelo"], corner_radius=0)
        self.topo.grid(row=0, column=0, sticky="ew")
        try:
            img_pil   = Image.open(LOGO_PATH)
            self.logo = ctk.CTkImage(light_image=img_pil, size=(110, 110))
            ctk.CTkLabel(self.topo, image=self.logo, text="").place(relx=0.5, rely=0.5, anchor="center")
        except Exception:
            ctk.CTkLabel(self.topo, text="DOM MOTA", font=("Segoe UI", 35, "bold"),
                         text_color=C["vermelho"]).place(relx=0.5, rely=0.5, anchor="center")

        self.container = ctk.CTkFrame(self, fg_color=C["fundo"], corner_radius=0)
        self.container.grid(row=1, column=0, sticky="nsew")
        ctk.CTkButton(self, text="ℹ Créditos", width=100, height=28,
                      fg_color="#555", hover_color="#333", corner_radius=8,
                      font=("Segoe UI", 11), command=self.mostrar_creditos
                      ).place(relx=1.0, rely=1.0, anchor="se", x=-16, y=-16)
        self.mostrar_tela_principal()

    def limpar_container(self):
        for w in self.container.winfo_children(): w.destroy()

    # ── MENU ─────────────────────────────────
    def mostrar_tela_principal(self):
        self.limpar_container()
        wrapper = ctk.CTkFrame(self.container, fg_color="transparent")
        wrapper.place(relx=0.5, rely=0.5, anchor="center")
        ctk.CTkLabel(wrapper, text="O que você deseja fazer?",
                     font=("Segoe UI", 15), text_color=C["subtexto"]).pack(pady=(0, 18))
        grid = ctk.CTkFrame(wrapper, fg_color="transparent"); grid.pack()
        icones = ["📋", "👤", "📚", "⏳", "🏆", "📊"]
        opcoes = [
            ("Empréstimos",      self.tela_emprestimo),
            ("Gerenciar Alunos", self.tela_alunos),
            ("Acervo",           self.tela_acervo),
            ("Pendências",       self.tela_pendentes),
            ("Ranking",          self.tela_ranking_unificado),
            ("Relatórios",       self.tela_relatorios),
        ]
        for i, ((txt, cmd), ico) in enumerate(zip(opcoes, icones)):
            card = ctk.CTkFrame(grid, fg_color=C["card"], corner_radius=16,
                                border_width=1, border_color=C["borda"], width=200, height=120)
            card.grid(row=i//3, column=i%3, padx=12, pady=12)
            card.grid_propagate(False)
            inner = ctk.CTkFrame(card, fg_color="transparent")
            inner.place(relx=0.5, rely=0.5, anchor="center")
            ctk.CTkLabel(inner, text=ico, font=("Segoe UI", 28)).pack()
            ctk.CTkLabel(inner, text=txt, font=("Segoe UI", 13, "bold"),
                         text_color=C["texto"]).pack(pady=(4, 8))
            ctk.CTkButton(inner, text="Abrir",
                          **estilo_btn_primario(width=100, height=30), command=cmd).pack()

    def mostrar_creditos(self):
        pop = ctk.CTkToplevel(self)
        pop.title("Créditos"); pop.geometry("420x280")
        pop.resizable(False, False); pop.attributes("-topmost", True)
        pop.configure(fg_color=C["card"])
        ctk.CTkLabel(pop, text="Sobre o Desenvolvedor",
                     font=FONTE_TITULO, text_color=C["vermelho"]).pack(pady=(24, 8))
        ctk.CTkFrame(pop, height=1, fg_color=C["borda"]).pack(fill="x", padx=30)
        ctk.CTkLabel(pop,
                     text="Davi Lucas Galdino Braz\nTécnico em Desenvolvimento de Sistemas\n\nProjeto orientado pela professora\nGislene Laurenço",
                     justify="center", font=FONTE_CORPO, text_color=C["texto"]).pack(pady=16)
        ctk.CTkButton(pop, text="Fechar", **estilo_btn_cinza(width=120), command=pop.destroy).pack(pady=8)

    # ── EMPRÉSTIMOS ───────────────────────────
    def tela_emprestimo(self):
        self.limpar_container()
        cabecalho_tela(self.container, "📋  Empréstimos", self.mostrar_tela_principal)
        centro = ctk.CTkFrame(self.container, fg_color=C["fundo"], corner_radius=0)
        centro.pack(fill="both", expand=True)

        abas_frame = ctk.CTkFrame(centro, fg_color="transparent")
        abas_frame.pack(fill="x", padx=30, pady=(16, 0))
        btn_nome  = ctk.CTkButton(abas_frame, text="🔍 Buscar por Nome",
                                   width=190, height=36, corner_radius=10,
                                   fg_color=C["vermelho"], hover_color=C["vermelho_esc"],
                                   text_color="white", font=FONTE_BOLD)
        btn_nome.pack(side="left", padx=(0, 8))
        btn_turma = ctk.CTkButton(abas_frame, text="🏫 Selecionar por Turma",
                                   width=210, height=36, corner_radius=10,
                                   fg_color=C["inativo"], hover_color="#D1D5DB",
                                   text_color=C["inativo_txt"], font=FONTE_CORPO)
        btn_turma.pack(side="left")

        area = ctk.CTkFrame(centro, fg_color="transparent")
        area.pack(fill="both", expand=True, padx=30, pady=12)

        def _montar_lista_alunos(parent, rows):
            for n, s, tr, tn in rows:
                card = ctk.CTkFrame(parent, fg_color=C["card"], corner_radius=10,
                                    border_width=1, border_color=C["borda"])
                card.pack(fill="x", pady=3)
                ctk.CTkLabel(card, text="👤", font=("Segoe UI", 18), width=44).pack(side="left", padx=10)
                info = ctk.CTkFrame(card, fg_color="transparent")
                info.pack(side="left", fill="x", expand=True, pady=10)
                ctk.CTkLabel(info, text=n, font=FONTE_BOLD, text_color=C["texto"], anchor="w").pack(anchor="w", padx=4)
                ctk.CTkLabel(info, text=f"Série: {s or '—'}  ·  Turma: {tr or '—'}  ·  Turno: {tn or '—'}",
                             font=FONTE_PEQUENA, text_color=C["subtexto"], anchor="w").pack(anchor="w", padx=4)
                ctk.CTkButton(card, text="Registrar Empréstimo", **estilo_btn_verde(width=180),
                              command=lambda x=n, y=s, z=tr, w=tn: self.pop_livro(x, y, z, w)
                              ).pack(side="right", padx=12, pady=10)

        def aba_nome():
            for w in area.winfo_children(): w.destroy()
            btn_nome.configure(fg_color=C["vermelho"], text_color="white", font=FONTE_BOLD)
            btn_turma.configure(fg_color=C["inativo"], text_color=C["inativo_txt"], font=FONTE_CORPO)
            card_busca = ctk.CTkFrame(area, fg_color=C["card"], corner_radius=14,
                                       border_width=1, border_color=C["borda"])
            card_busca.pack(fill="x")
            ctk.CTkLabel(card_busca, text="Buscar aluno por nome",
                         font=FONTE_SECAO, text_color=C["vermelho"]).pack(anchor="w", padx=20, pady=(16, 8))
            linha = ctk.CTkFrame(card_busca, fg_color="transparent")
            linha.pack(fill="x", padx=20, pady=(0, 16))
            busc = ctk.CTkEntry(linha, placeholder_text="🔍 Digite o nome do aluno...", **estilo_entry(500))
            busc.pack(side="left"); busc.focus_set()
            sc = ctk.CTkScrollableFrame(area, fg_color="transparent", corner_radius=0)
            sc.pack(fill="both", expand=True, pady=(8, 0))
            _d = [None]
            def query(e=None):
                if _d[0]: self.after_cancel(_d[0])
                _d[0] = self.after(700, _exec)
            def _exec():
                t = busc.get().upper().strip()
                if len(t) < 2:
                    for w in sc.winfo_children(): w.destroy()
                    return
                def buscar():
                    with _conectar() as conn:
                        c = conn.cursor()
                        c.execute("SELECT nome, serie, turma, turno FROM alunos WHERE nome LIKE ? COLLATE NOCASE LIMIT 15", (f"%{t}%",))
                        return c.fetchall()
                def mostrar(rows):
                    if not sc.winfo_exists():
                        return
                    for w in sc.winfo_children(): w.destroy()
                    if not rows:
                        ctk.CTkLabel(sc, text="Nenhum aluno encontrado.", font=FONTE_CORPO, text_color=C["subtexto"]).pack(pady=20)
                        return
                    _montar_lista_alunos(sc, rows)
                em_thread(buscar, mostrar)
            busc.bind("<KeyRelease>", query)

        def aba_turma():
            for w in area.winfo_children(): w.destroy()
            btn_turma.configure(fg_color=C["vermelho"], text_color="white", font=FONTE_BOLD)
            btn_nome.configure(fg_color=C["inativo"], text_color=C["inativo_txt"], font=FONTE_CORPO)
            split = ctk.CTkFrame(area, fg_color="transparent")
            split.pack(fill="both", expand=True)
            split.columnconfigure(0, weight=0); split.columnconfigure(1, weight=1); split.rowconfigure(0, weight=1)
            panel_esq = ctk.CTkFrame(split, fg_color=C["card"], corner_radius=14,
                                      border_width=1, border_color=C["borda"], width=180)
            panel_esq.grid(row=0, column=0, sticky="nsew", padx=(0, 10)); panel_esq.grid_propagate(False)
            ctk.CTkLabel(panel_esq, text="Turmas", font=FONTE_SECAO, text_color=C["vermelho"]).pack(anchor="w", padx=14, pady=(14, 6))
            separador(panel_esq)
            sc_turmas = ctk.CTkScrollableFrame(panel_esq, fg_color="transparent", corner_radius=0, width=160)
            sc_turmas.pack(fill="both", expand=True, padx=4, pady=4)
            panel_dir = ctk.CTkFrame(split, fg_color="transparent")
            panel_dir.grid(row=0, column=1, sticky="nsew")
            lbl_sel = ctk.CTkLabel(panel_dir, text="← Selecione uma turma", font=FONTE_CORPO, text_color=C["subtexto"])
            lbl_sel.pack(anchor="w", pady=(6, 4))
            sc_alunos = ctk.CTkScrollableFrame(panel_dir, fg_color="transparent", corner_radius=0)
            sc_alunos.pack(fill="both", expand=True)
            botoes = []
            def abrir_turma(s, t, tn, btn_clicado):
                for b in botoes: b.configure(fg_color=C["fundo"], text_color=C["texto"])
                btn_clicado.configure(fg_color=C["vermelho"], text_color="white")
                lbl_sel.configure(text=f"Alunos — {s}º ano  Turma {t}  {tn}", text_color=C["vermelho"], font=FONTE_BOLD)
                def buscar():
                    with _conectar() as conn:
                        c = conn.cursor()
                        c.execute("SELECT nome, serie, turma, turno FROM alunos WHERE serie=? AND turma=? AND turno=? ORDER BY nome", (s, t, tn))
                        return c.fetchall()
                def mostrar(rows):
                    for w in sc_alunos.winfo_children(): w.destroy()
                    _montar_lista_alunos(sc_alunos, rows)
                em_thread(buscar, mostrar)
            def carregar_turmas():
                with _conectar() as conn:
                    c = conn.cursor()
                    c.execute("SELECT DISTINCT serie, turma, turno FROM alunos WHERE serie!='' AND turma!='' AND turno!='' ORDER BY CAST(serie AS INTEGER), turma, turno")
                    return c.fetchall()
            def popular(turmas_unicas):
                for s, t, tn in turmas_unicas:
                    turno_curto = "M" if "MAN" in tn.upper() else "T" if "TAR" in tn.upper() else tn[:1]
                    label = f"{s}º  T{t}  {turno_curto}"
                    btn = ctk.CTkButton(sc_turmas, text=label, width=148, height=40, corner_radius=8,
                                         fg_color=C["fundo"], hover_color=C["borda"], text_color=C["texto"],
                                         font=FONTE_BOLD, border_width=1, border_color=C["borda"], anchor="w")
                    btn.configure(command=lambda ss=s, tt=t, ttn=tn, b=btn: abrir_turma(ss, tt, ttn, b))
                    btn.pack(fill="x", pady=2); botoes.append(btn)
            em_thread(carregar_turmas, popular)

        btn_nome.configure(command=aba_nome)
        btn_turma.configure(command=aba_turma)
        aba_nome()

    def pop_livro(self, aluno, serie, turma, turno):
        pop = ctk.CTkToplevel(self)
        pop.geometry("480x380"); pop.resizable(False, False)
        pop.attributes("-topmost", True); pop.title("Registrar Empréstimo")
        pop.configure(fg_color=C["card"])
        ctk.CTkLabel(pop, text="📋  Novo Empréstimo", font=FONTE_TITULO, text_color=C["vermelho"]).pack(pady=(24, 4))
        ctk.CTkLabel(pop, text=f"Aluno: {aluno}", font=FONTE_BOLD, text_color=C["texto"]).pack()
        ctk.CTkFrame(pop, height=1, fg_color=C["borda"]).pack(fill="x", padx=30, pady=16)
        f_cod = ctk.CTkFrame(pop, fg_color="transparent"); f_cod.pack()
        lbl_sub(f_cod, "Código:"); cod = ctk.CTkEntry(f_cod, placeholder_text="0000", **estilo_entry(80))
        cod.pack(side="left", padx=(0, 16))
        l = ctk.CTkEntry(pop, placeholder_text="Título do Livro", **estilo_entry(380))
        d = ctk.CTkEntry(pop, placeholder_text="Data de Entrega  (ex: 15/08/2025)", **estilo_entry(380))
        l.pack(pady=6); d.pack(pady=6)
        def salvar():
            titulo_raw = l.get().strip()
            if not titulo_raw:
                messagebox.showwarning("Atenção", "Informe o título.", parent=pop); return
            titulo = (f"[{cod.get().strip()}] {titulo_raw.upper()}" if cod.get().strip() else titulo_raw.upper())
            try:
                with _conectar() as conn:
                    conn.execute("INSERT INTO emprestimos (aluno_nome, livro_titulo, data_entrega, serie, turma, turno) VALUES (?,?,?,?,?,?)",
                                 (aluno, titulo, d.get().strip(), serie, turma, turno))
                    conn.commit()
                pop.destroy()
                messagebox.showinfo("Sucesso", f"Empréstimo registrado para {aluno}!")
            except sqlite3.Error as e:
                messagebox.showerror("Erro", f"Não foi possível registrar.\n\n{e}", parent=pop)
        ctk.CTkButton(pop, text="✓  Confirmar Empréstimo", **estilo_btn_verde(240, 42), command=salvar).pack(pady=20)

    # ── GERENCIAR ALUNOS ──────────────────────
    def tela_alunos(self):
        self.limpar_container()
        cabecalho_tela(self.container, "👤  Gerenciar Alunos", self.mostrar_tela_principal)
        centro = ctk.CTkFrame(self.container, fg_color=C["fundo"], corner_radius=0)
        centro.pack(fill="both", expand=True)
        card_add = ctk.CTkFrame(centro, fg_color=C["card"], corner_radius=14, border_width=1, border_color=C["borda"])
        card_add.pack(fill="x", padx=30, pady=(20, 0))
        ctk.CTkLabel(card_add, text="Cadastrar aluno", font=FONTE_SECAO, text_color=C["vermelho"]).pack(anchor="w", padx=20, pady=(14, 8))
        linha = ctk.CTkFrame(card_add, fg_color="transparent"); linha.pack(fill="x", padx=20, pady=(0, 16))
        en = ctk.CTkEntry(linha, placeholder_text="Nome completo", **estilo_entry(220)); en.pack(side="left", padx=(0, 8))
        lbl_sub(linha, "Série:")
        es = ctk.CTkEntry(linha, placeholder_text="Ex: 6", **estilo_entry(70)); es.pack(side="left", padx=(0, 8))
        lbl_sub(linha, "Turma:")
        seg_turma = ctk.CTkSegmentedButton(linha, values=TURMAS_OPCOES, **estilo_segmented(160))
        seg_turma.set("1"); seg_turma.pack(side="left", padx=(0, 8))
        def atualizar_turmas(e=None):
            s = es.get().strip().upper()
            opcoes = turmas_da_serie(s) if s else TURMAS_OPCOES
            seg_turma.configure(values=opcoes); seg_turma.set(opcoes[0])
        es.bind("<FocusOut>", atualizar_turmas); es.bind("<KeyRelease>", atualizar_turmas)
        lbl_sub(linha, "Turno:")
        etu = ctk.CTkEntry(linha, placeholder_text="Manhã / Tarde", **estilo_entry(130)); etu.pack(side="left", padx=(0, 12))
        def salvar():
            if en.get().strip():
                nome_up = en.get().upper().strip(); serie = es.get().upper().strip()
                turma = seg_turma.get().strip(); turno = etu.get().upper().strip()
                try:
                    with _conectar() as conn:
                        conn.execute("INSERT INTO alunos (nome, serie, turma, turno) VALUES (?,?,?,?)", (nome_up, serie, turma, turno))
                        if serie and turma and turno:
                            conn.execute("INSERT OR IGNORE INTO turmas (serie, turma, turno) VALUES (?,?,?)", (serie, turma, turno))
                        conn.commit()
                    en.delete(0, "end"); aba_nome()
                except sqlite3.Error as e:
                    messagebox.showerror("Erro", str(e))
        ctk.CTkButton(linha, text="+ Cadastrar", **estilo_btn_verde(140), command=salvar).pack(side="left")

        abas = ctk.CTkFrame(centro, fg_color="transparent"); abas.pack(fill="x", padx=30, pady=(10, 0))
        btn_nome_a = ctk.CTkButton(abas, text="🔍 Buscar por Nome", width=190, height=38, corner_radius=10,
                                    fg_color=C["vermelho"], hover_color=C["vermelho_esc"], text_color="white", font=FONTE_BOLD)
        btn_nome_a.pack(side="left", padx=(0, 8))
        btn_turma_a = ctk.CTkButton(abas, text="🏫 Selecionar por Turma", width=210, height=38, corner_radius=10,
                                     fg_color=C["inativo"], hover_color="#D1D5DB", text_color=C["inativo_txt"], font=FONTE_CORPO)
        btn_turma_a.pack(side="left")
        area = ctk.CTkFrame(centro, fg_color="transparent"); area.pack(fill="both", expand=True, padx=30, pady=8)

        def montar_lista(pai, rows):
            for nome, serie, turma, turno in rows:
                card = ctk.CTkFrame(pai, fg_color=C["card"], corner_radius=10, border_width=1, border_color=C["borda"])
                card.pack(fill="x", pady=3)
                ctk.CTkLabel(card, text="👤", font=("Segoe UI", 16), width=40).pack(side="left", padx=10)
                info = ctk.CTkFrame(card, fg_color="transparent"); info.pack(side="left", fill="x", expand=True, pady=8)
                ctk.CTkLabel(info, text=nome, font=FONTE_BOLD, text_color=C["texto"], anchor="w").pack(anchor="w", padx=4)
                ctk.CTkLabel(info, text=f"Série: {serie or '—'}  ·  Turma: {turma or '—'}  ·  Turno: {turno or '—'}",
                             font=FONTE_PEQUENA, text_color=C["subtexto"], anchor="w").pack(anchor="w", padx=4)
                ctk.CTkButton(card, text="✕", width=32, height=32, fg_color="#FFEEEE", hover_color="#FFCCCC",
                              text_color=C["vermelho"], corner_radius=8, font=FONTE_BOLD,
                              command=lambda x=nome: deletar(x)).pack(side="right", padx=12)

        def deletar(nome):
            if messagebox.askyesno("Confirmar", f"Excluir o aluno '{nome}'?"):
                try:
                    with _conectar() as conn:
                        conn.execute("DELETE FROM alunos WHERE nome=?", (nome,)); conn.commit()
                    aba_nome()
                except sqlite3.Error as e:
                    messagebox.showerror("Erro", str(e))

        def aba_nome():
            for w in area.winfo_children(): w.destroy()
            btn_nome_a.configure(fg_color=C["vermelho"], text_color="white", font=FONTE_BOLD)
            btn_turma_a.configure(fg_color=C["inativo"], text_color=C["inativo_txt"], font=FONTE_CORPO)
            card = ctk.CTkFrame(area, fg_color=C["card"], corner_radius=14, border_width=1, border_color=C["borda"])
            card.pack(fill="x")
            ctk.CTkLabel(card, text="Buscar aluno por nome", font=FONTE_SECAO, text_color=C["vermelho"]).pack(anchor="w", padx=20, pady=(16, 8))
            linha2 = ctk.CTkFrame(card, fg_color="transparent"); linha2.pack(fill="x", padx=20, pady=(0, 16))
            busc = ctk.CTkEntry(linha2, placeholder_text="🔍 Pesquisar...", **estilo_entry(420)); busc.pack(side="left")
            sc = ctk.CTkScrollableFrame(area, fg_color="transparent"); sc.pack(fill="both", expand=True, pady=(10, 0))
            _d = [None]
            def pesquisar(e=None):
                if _d[0]: self.after_cancel(_d[0])
                _d[0] = self.after(700, _exec)
            def _exec():
                texto = busc.get().upper().strip()
                def buscar():
                    with _conectar() as conn:
                        c = conn.cursor()
                        if texto:
                            c.execute("SELECT nome, serie, turma, turno FROM alunos WHERE nome LIKE ? COLLATE NOCASE ORDER BY nome LIMIT 100", (f"%{texto}%",))
                        else:
                            c.execute("SELECT nome, serie, turma, turno FROM alunos ORDER BY nome LIMIT 30")
                        return c.fetchall()
                def mostrar(rows):
                    if not sc.winfo_exists():
                        return
                    for w in sc.winfo_children(): w.destroy()
                    montar_lista(sc, rows)
                em_thread(buscar, mostrar)
            busc.bind("<KeyRelease>", pesquisar); _exec()

        def aba_turma_alunos():
            for w in area.winfo_children(): w.destroy()
            btn_turma_a.configure(fg_color=C["vermelho"], text_color="white", font=FONTE_BOLD)
            btn_nome_a.configure(fg_color=C["inativo"], text_color=C["inativo_txt"], font=FONTE_CORPO)
            split = ctk.CTkFrame(area, fg_color="transparent"); split.pack(fill="both", expand=True)
            esq = ctk.CTkFrame(split, fg_color=C["card"], corner_radius=14, border_width=1, border_color=C["borda"], width=180)
            esq.pack(side="left", fill="y", padx=(0, 10))
            ctk.CTkLabel(esq, text="Turmas", font=FONTE_SECAO, text_color=C["vermelho"]).pack(anchor="w", padx=14, pady=(14, 8))
            sc_t = ctk.CTkScrollableFrame(esq, fg_color="transparent", width=160); sc_t.pack(fill="both", expand=True, padx=4, pady=4)
            dir_ = ctk.CTkFrame(split, fg_color="transparent"); dir_.pack(side="left", fill="both", expand=True)
            titulo = ctk.CTkLabel(dir_, text="← Selecione uma turma", font=FONTE_CORPO, text_color=C["subtexto"])
            titulo.pack(anchor="w", pady=(6, 4))
            sc_a = ctk.CTkScrollableFrame(dir_, fg_color="transparent"); sc_a.pack(fill="both", expand=True)
            botoes = []
            def abrir(s, t, tn, b_clicado):
                for b in botoes: b.configure(fg_color=C["fundo"], text_color=C["texto"])
                b_clicado.configure(fg_color=C["vermelho"], text_color="white")
                titulo.configure(text=f"{s}º Ano • Turma {t} • {tn}", text_color=C["vermelho"])
                def buscar():
                    with _conectar() as conn:
                        c = conn.cursor()
                        c.execute("SELECT nome, serie, turma, turno FROM alunos WHERE serie=? AND turma=? AND turno=? ORDER BY nome", (s, t, tn))
                        return c.fetchall()
                def mostrar(rows):
                    for w in sc_a.winfo_children(): w.destroy()
                    montar_lista(sc_a, rows)
                em_thread(buscar, mostrar)
            def carregar():
                with _conectar() as conn:
                    c = conn.cursor()
                    c.execute("SELECT DISTINCT serie, turma, turno FROM alunos ORDER BY serie, turma, turno")
                    return c.fetchall()
            def popular(turmas):
                for s, t, tn in turmas:
                    btn = ctk.CTkButton(sc_t, text=f"{s}º  T{t}", width=148, height=40, corner_radius=8,
                                        fg_color=C["fundo"], hover_color=C["borda"], text_color=C["texto"],
                                        font=FONTE_BOLD, border_width=1, border_color=C["borda"], anchor="w")
                    btn.configure(command=lambda ss=s, tt=t, ttn=tn, b=btn: abrir(ss, tt, ttn, b))
                    btn.pack(fill="x", pady=2); botoes.append(btn)
            em_thread(carregar, popular)

        btn_nome_a.configure(command=aba_nome)
        btn_turma_a.configure(command=aba_turma_alunos)
        aba_nome()

    # ── ACERVO ───────────────────────────────
    # CORREÇÃO PRINCIPAL: detecta dinamicamente as colunas do banco
    def tela_acervo(self):
        self.limpar_container()
        cabecalho_tela(self.container, "📚  Acervo de Livros", self.mostrar_tela_principal)
        centro = ctk.CTkFrame(self.container, fg_color=C["fundo"], corner_radius=0)
        centro.pack(fill="both", expand=True)

        # Detecta quais colunas existem no banco para acervo
        colunas_db = _colunas_acervo()
        tem_edicao  = "edição"  in colunas_db
        tem_estante = "estante" in colunas_db
        tem_fileira = "fileira" in colunas_db
        tem_codigo  = "codigo"  in colunas_db

        # ── Card adicionar ───────────────────────────────────────────
        card_add = ctk.CTkFrame(centro, fg_color=C["card"], corner_radius=14, border_width=1, border_color=C["borda"])
        card_add.pack(fill="x", padx=30, pady=(20, 0))
        ctk.CTkLabel(card_add, text="Adicionar livro", font=FONTE_SECAO, text_color=C["vermelho"]).pack(anchor="w", padx=20, pady=(14, 6))

        l1 = ctk.CTkFrame(card_add, fg_color="transparent"); l1.pack(fill="x", padx=20, pady=(0, 4))
        et = ctk.CTkEntry(l1, placeholder_text="Título",  **estilo_entry(260))
        ea = ctk.CTkEntry(l1, placeholder_text="Autor",   **estilo_entry(200))
        eg = ctk.CTkEntry(l1, placeholder_text="Gênero",  **estilo_entry(140))
        for w in [et, ea, eg]: w.pack(side="left", padx=(0, 8))

        l2 = ctk.CTkFrame(card_add, fg_color="transparent"); l2.pack(fill="x", padx=20, pady=(0, 14))

        # Edição (se existe no banco)
        eed = None
        if tem_edicao:
            lbl_sub(l2, "Edição:")
            eed = ctk.CTkEntry(l2, placeholder_text="Ex: 1ª", **estilo_entry(90)); eed.pack(side="left", padx=(0, 16))

        # Código
        ecod = None
        if tem_codigo:
            lbl_sub(l2, "Código (4 dígitos):")
            ecod = ctk.CTkEntry(l2, placeholder_text="0000", **estilo_entry(90)); ecod.pack(side="left", padx=(0, 16))

        # Estante
        seg_est = None
        if tem_estante:
            lbl_sub(l2, "Estante:")
            seg_est = ctk.CTkSegmentedButton(l2, values=ESTANTES_LETRAS, **estilo_segmented(400))
            seg_est.set("A"); seg_est.pack(side="left", padx=(0, 16))

        # Fileira
        seg_fil = None
        if tem_fileira:
            lbl_sub(l2, "Fileira:")
            seg_fil = ctk.CTkSegmentedButton(l2, values=FILEIRAS_NUM, **estilo_segmented(200))
            seg_fil.set("1"); seg_fil.pack(side="left", padx=(0, 16))

        ctk.CTkButton(l2, text="+ Adicionar", **estilo_btn_verde(130), command=lambda: add()).pack(side="left")

        # ── Card pesquisar ───────────────────────────────────────────
        card_busc = ctk.CTkFrame(centro, fg_color=C["card"], corner_radius=14, border_width=1, border_color=C["borda"])
        card_busc.pack(fill="x", padx=30, pady=(10, 0))
        ctk.CTkLabel(card_busc, text="Pesquisar livros", font=FONTE_SECAO, text_color=C["vermelho"]).pack(anchor="w", padx=20, pady=(14, 6))

        lb = ctk.CTkFrame(card_busc, fg_color="transparent"); lb.pack(fill="x", padx=20, pady=(0, 14))
        bt = ctk.CTkEntry(lb, placeholder_text="🔍 Título",  **estilo_entry(220)); bt.pack(side="left", padx=(0, 8))
        ba = ctk.CTkEntry(lb, placeholder_text="👤 Autor",   **estilo_entry(160)); ba.pack(side="left", padx=(0, 8))
        bg = ctk.CTkEntry(lb, placeholder_text="🏷 Gênero",  **estilo_entry(120)); bg.pack(side="left", padx=(0, 8))

        bed = None
        if tem_edicao:
            bed = ctk.CTkEntry(lb, placeholder_text="📖 Edição", **estilo_entry(100)); bed.pack(side="left", padx=(0, 8))

        bcd = None
        if tem_codigo:
            bcd = ctk.CTkEntry(lb, placeholder_text="# Código", **estilo_entry(90)); bcd.pack(side="left", padx=(0, 8))

        be = None
        if tem_estante:
            be = ctk.CTkEntry(lb, placeholder_text="📦 Estante", **estilo_entry(100)); be.pack(side="left", padx=(0, 8))

        bf = None
        if tem_fileira:
            bf = ctk.CTkEntry(lb, placeholder_text="📍 Fileira", **estilo_entry(90)); bf.pack(side="left", padx=(0, 8))

        campos_busca = [w for w in [bt, ba, bg, bed, bcd, be, bf] if w is not None]
        ctk.CTkButton(lb, text="Limpar", **estilo_btn_cinza(90), command=lambda: limpar_busca()).pack(side="left")

        lbl_count = ctk.CTkLabel(centro, text="", font=FONTE_PEQUENA, text_color=C["subtexto"])
        lbl_count.pack(anchor="w", padx=34, pady=(8, 2))
        sc = ctk.CTkScrollableFrame(centro, fg_color="transparent", corner_radius=0)
        sc.pack(fill="both", expand=True, padx=30, pady=(4, 20))

        _d = [None]

        def listar(e=None):
            if _d[0]: self.after_cancel(_d[0])
            _d[0] = self.after(700, _executar_listar)

        def _executar_listar():
            qt  = bt.get().upper().strip()
            qa  = ba.get().upper().strip()
            qg  = bg.get().upper().strip()
            qed = bed.get().strip() if bed else ""
            qcd = bcd.get().strip() if bcd else ""
            qe  = be.get().upper().strip() if be else ""
            qf  = bf.get().strip() if bf else ""

            # Monta SELECT com as colunas que realmente existem
            colunas_sel = ["titulo", "autor", "genero"]
            if tem_edicao:  colunas_sel.append('"edição"')
            if tem_codigo:  colunas_sel.append("codigo")
            if tem_estante: colunas_sel.append("estante")
            if tem_fileira: colunas_sel.append("fileira")

            sql = f"SELECT {', '.join(colunas_sel)} FROM acervo WHERE 1=1"
            params = []
            if qt:  sql += " AND titulo  LIKE ? COLLATE NOCASE"; params.append(f"%{qt}%")
            if qa:  sql += " AND autor   LIKE ? COLLATE NOCASE"; params.append(f"%{qa}%")
            if qg:  sql += " AND genero  LIKE ? COLLATE NOCASE"; params.append(f"%{qg}%")
            if qed and tem_edicao:  sql += ' AND "edição" LIKE ?'; params.append(f"%{qed}%")
            if qcd and tem_codigo:  sql += " AND codigo  LIKE ?"; params.append(f"%{qcd}%")
            if qe  and tem_estante: sql += " AND estante LIKE ?"; params.append(f"%{qe}%")
            if qf  and tem_fileira: sql += " AND fileira LIKE ?"; params.append(f"%{qf}%")

            sem_filtro = not any([qt, qa, qg, qed, qcd, qe, qf])
            sql += " ORDER BY titulo" + (" LIMIT 30" if sem_filtro else " LIMIT 80")

            params_snap = list(params)

            def buscar():
                with _conectar() as conn:
                    c = conn.cursor()
                    c.execute(sql, params_snap)
                    return c.fetchall(), sem_filtro

            def mostrar(resultado):
                if not sc.winfo_exists():
                    return
                rows, sf = resultado
                if len(sc.winfo_children()) > 250:
                    for w in sc.winfo_children(): w.destroy()
                else:
                    for w in sc.winfo_children(): w.destroy()
                sufixo = " — primeiros 30" if sf and len(rows) == 30 else ""
                lbl_count.configure(text=f"{len(rows)} livro(s) encontrado(s){sufixo}")

                for i, row in enumerate(rows):
                    if i > 80:
                        break
                    # Mapeia posições dinamicamente
                    idx = 0
                    t_val = row[idx]; idx += 1
                    a_val = row[idx]; idx += 1
                    g_val = row[idx]; idx += 1
                    ed_val  = row[idx] if tem_edicao  else None; idx += (1 if tem_edicao else 0)
                    cod_val = row[idx] if tem_codigo  else None; idx += (1 if tem_codigo else 0)
                    est_val = row[idx] if tem_estante else None; idx += (1 if tem_estante else 0)
                    fil_val = row[idx] if tem_fileira else None

                    card = ctk.CTkFrame(sc, fg_color=C["card"], corner_radius=10,
                                        border_width=1, border_color=C["borda"])
                    card.pack(fill="x", pady=3)

                    # Badge código (se existir) ou ícone
                    badge = cod_val if (cod_val and cod_val.strip()) else "📖"
                    ctk.CTkLabel(card, text=badge, font=("Courier", 13, "bold"),
                                 text_color=C["vermelho"], width=52).pack(side="left", padx=(10, 0))

                    info = ctk.CTkFrame(card, fg_color="transparent")
                    info.pack(side="left", fill="x", expand=True, pady=8)
                    ctk.CTkLabel(info, text=t_val if len(t_val) <= 60 else t_val[:57]+"...",
                                 font=FONTE_BOLD, text_color=C["texto"], anchor="w").pack(anchor="w", padx=4)
                    partes = []
                    if a_val and a_val.strip():   partes.append(f"Autor: {a_val}")
                    if g_val and g_val.strip():   partes.append(f"Gênero: {g_val}")
                    if ed_val and ed_val.strip(): partes.append(f"Edição: {ed_val}")
                    if est_val and est_val.strip(): partes.append(f"Estante: {est_val}")
                    if fil_val and fil_val.strip(): partes.append(f"Fileira: {fil_val}")
                    if partes:
                        ctk.CTkLabel(info, text="  ·  ".join(partes), font=FONTE_PEQUENA,
                                     text_color=C["subtexto"], anchor="w").pack(anchor="w", padx=4)
                    ctk.CTkButton(card, text="✕", width=32, height=32,
                                  fg_color="#FFEEEE", hover_color="#FFCCCC",
                                  text_color=C["vermelho"], corner_radius=8, font=FONTE_BOLD,
                                  command=lambda x=t_val: deletar_l(x)).pack(side="right", padx=12)

            em_thread(buscar, mostrar)

        def add():
            titulo_raw = et.get().strip()
            if not titulo_raw:
                messagebox.showwarning("Atenção", "Informe o título do livro."); return
            cod_val = ecod.get().strip() if ecod else ""
            if cod_val and (not cod_val.isdigit() or len(cod_val) != 4):
                messagebox.showwarning("Atenção", "O código deve ter exatamente 4 dígitos numéricos."); return

            # Monta INSERT com as colunas que existem
            colunas_ins = ["titulo", "autor", "genero"]
            valores_ins = [titulo_raw.upper(), ea.get().upper().strip(), eg.get().upper().strip()]
            if tem_edicao and eed:
                colunas_ins.append('"edição"'); valores_ins.append(eed.get().strip())
            if tem_codigo:
                colunas_ins.append("codigo"); valores_ins.append(cod_val)
            if tem_estante and seg_est:
                colunas_ins.append("estante"); valores_ins.append(seg_est.get().strip())
            if tem_fileira and seg_fil:
                colunas_ins.append("fileira"); valores_ins.append(seg_fil.get().strip())

            placeholders = ", ".join(["?"] * len(valores_ins))
            sql_ins = f"INSERT INTO acervo ({', '.join(colunas_ins)}) VALUES ({placeholders})"
            try:
                with _conectar() as conn:
                    conn.execute(sql_ins, valores_ins); conn.commit()
                for w in [et, ea, eg] + ([eed] if eed else []) + ([ecod] if ecod else []):
                    if w: w.delete(0, "end")
                listar()
            except sqlite3.Error as e:
                messagebox.showerror("Erro ao adicionar", f"Não foi possível adicionar o livro.\n\n{e}")

        def deletar_l(t):
            if messagebox.askyesno("Confirmar", f"Excluir '{t}' do acervo?"):
                try:
                    with _conectar() as conn:
                        conn.execute("DELETE FROM acervo WHERE titulo=?", (t,)); conn.commit()
                    listar()
                except sqlite3.Error as e:
                    messagebox.showerror("Erro ao excluir", str(e))

        def limpar_busca():
            for w in campos_busca: w.delete(0, "end")
            listar()

        for w in campos_busca: w.bind("<KeyRelease>", listar)
        listar()

    # ── PENDÊNCIAS ───────────────────────────
    def tela_pendentes(self):
        self.limpar_container()
        cabecalho_tela(self.container, "⏳  Pendências de Devolução", self.mostrar_tela_principal)
        centro = ctk.CTkFrame(self.container, fg_color=C["fundo"], corner_radius=0)
        centro.pack(fill="both", expand=True)
        lbl_count = ctk.CTkLabel(centro, text="", font=FONTE_PEQUENA, text_color=C["subtexto"])
        lbl_count.pack(anchor="w", padx=34, pady=(12, 4))
        sc = ctk.CTkScrollableFrame(centro, fg_color="transparent", corner_radius=0)
        sc.pack(fill="both", expand=True, padx=30, pady=(4, 20))
        def carregar():
            def buscar():
                with _conectar() as conn:
                    c = conn.cursor()
                    c.execute("SELECT rowid, aluno_nome, livro_titulo, data_entrega, serie, turma, turno FROM emprestimos ORDER BY aluno_nome")
                    return c.fetchall()
            def mostrar(rows):
                for w in sc.winfo_children(): w.destroy()
                lbl_count.configure(text=f"{len(rows)} empréstimo(s) pendente(s)")
                if not rows:
                    ctk.CTkLabel(sc, text="✅  Nenhuma pendência no momento.", font=FONTE_CORPO, text_color=C["subtexto"]).pack(pady=30)
                    return
                for rid, n, l, dt, s, tr, tn in rows:
                    card = ctk.CTkFrame(sc, fg_color=C["card"], corner_radius=10, border_width=1, border_color=C["borda"])
                    card.pack(fill="x", pady=3)
                    ctk.CTkLabel(card, text="📖", font=("Segoe UI", 18), width=42).pack(side="left", padx=10)
                    info = ctk.CTkFrame(card, fg_color="transparent"); info.pack(side="left", fill="x", expand=True, pady=8)
                    ctk.CTkLabel(info, text=l if len(l) <= 55 else l[:52]+"...", font=FONTE_BOLD, text_color=C["texto"], anchor="w").pack(anchor="w", padx=4)
                    meta = f"Aluno: {n}  ·  Série: {s or '—'}  ·  Turma: {tr or '—'}"
                    if dt: meta += f"  ·  Entrega: {dt}"
                    ctk.CTkLabel(info, text=meta, font=FONTE_PEQUENA, text_color=C["subtexto"], anchor="w").pack(anchor="w", padx=4)
                    ctk.CTkButton(card, text="✓ Devolver", **estilo_btn_verde(120),
                                  command=lambda x=rid, y=n, sa=s, ta=tr, tu=tn: self.finalizar(x, y, sa, ta, tu, carregar)
                                  ).pack(side="right", padx=12, pady=10)
            em_thread(buscar, mostrar)
        carregar()

    def finalizar(self, rid, nome, serie, turma, turno, cb):
        try:
            with _conectar() as conn:
                conn.execute("INSERT INTO ranking (nome, lidos, serie, turma, turno) VALUES (?,1,?,?,?) ON CONFLICT(nome) DO UPDATE SET lidos=lidos+1",
                             (nome, serie, turma, turno))
                conn.execute("DELETE FROM emprestimos WHERE rowid=?", (rid,))
                conn.commit()
            cb()
        except sqlite3.Error as e:
            messagebox.showerror("Erro ao devolver", str(e))

    # ── RANKING ──────────────────────────────
    def tela_ranking_unificado(self):
        self.limpar_container()
        cabecalho_tela(self.container, "🏆  Ranking de Leitura", self.mostrar_tela_principal)
        centro = ctk.CTkFrame(self.container, fg_color=C["fundo"], corner_radius=0)
        centro.pack(fill="both", expand=True)
        barra = ctk.CTkFrame(centro, fg_color=C["card"], corner_radius=0, height=56)
        barra.pack(fill="x"); barra.pack_propagate(False)
        ctk.CTkButton(barra, text="📥 Exportar Excel", **estilo_btn_verde(160), command=self.exportar_ranking).pack(side="left", padx=20, pady=9)
        ctk.CTkButton(barra, text="🗑 Zerar Ranking", width=150, height=38, corner_radius=10,
                      fg_color=C["perigo"], hover_color=C["perigo_hover"], text_color="white", font=FONTE_BOLD,
                      command=self.zerar_ranking).pack(side="left", pady=9)
        abas = ctk.CTkFrame(centro, fg_color="transparent"); abas.pack(fill="x", padx=30, pady=(16, 0))
        btn_alunos = ctk.CTkButton(abas, text="🏆 Alunos", width=160, height=38, corner_radius=10,
                                   fg_color=C["vermelho"], hover_color=C["vermelho_esc"], text_color="white", font=FONTE_BOLD)
        btn_alunos.pack(side="left", padx=(0, 8))
        btn_turmas = ctk.CTkButton(abas, text="🏫 Turmas", width=160, height=38, corner_radius=10,
                                   fg_color=C["inativo"], hover_color="#D1D5DB", text_color=C["inativo_txt"], font=FONTE_CORPO)
        btn_turmas.pack(side="left")
        sc = ctk.CTkScrollableFrame(centro, fg_color="transparent", corner_radius=0)
        sc.pack(fill="both", expand=True, padx=30, pady=12)
        medalhas = {1:"🥇", 2:"🥈", 3:"🥉"}

        def renderizar_alunos():
            btn_alunos.configure(fg_color=C["vermelho"], text_color="white", font=FONTE_BOLD)
            btn_turmas.configure(fg_color=C["inativo"], text_color=C["inativo_txt"], font=FONTE_CORPO)
            def buscar():
                with _conectar() as conn:
                    c = conn.cursor()
                    c.execute("SELECT nome, lidos, serie, turma, turno FROM ranking WHERE lidos > 0 ORDER BY lidos DESC")
                    return c.fetchall()
            def mostrar(rows):
                for w in sc.winfo_children(): w.destroy()
                if not rows:
                    ctk.CTkLabel(sc, text="Nenhum aluno no ranking ainda.", font=FONTE_CORPO, text_color=C["subtexto"]).pack(pady=40); return
                for i, (n, l, s, tr, tn) in enumerate(rows, 1):
                    card = ctk.CTkFrame(sc, fg_color=C["card"], corner_radius=10, border_width=1, border_color=C["borda"])
                    card.pack(fill="x", pady=3)
                    ctk.CTkLabel(card, text=medalhas.get(i, f"{i}º"), font=("Segoe UI", 20), width=52).pack(side="left", padx=8)
                    info = ctk.CTkFrame(card, fg_color="transparent"); info.pack(side="left", fill="x", expand=True, pady=10)
                    ctk.CTkLabel(info, text=n, font=FONTE_BOLD, text_color=C["texto"], anchor="w").pack(anchor="w", padx=4)
                    ctk.CTkLabel(info, text=f"Série: {s or '—'}  ·  Turma: {tr or '—'}  ·  Turno: {tn or '—'}",
                                 font=FONTE_PEQUENA, text_color=C["subtexto"], anchor="w").pack(anchor="w", padx=4)
                    ctk.CTkLabel(card, text=f"{l} livro{'s' if l!=1 else ''}", font=FONTE_BOLD, text_color=C["verde"], width=90).pack(side="right", padx=16)
            em_thread(buscar, mostrar)

        def renderizar_turmas():
            btn_turmas.configure(fg_color=C["vermelho"], text_color="white", font=FONTE_BOLD)
            btn_alunos.configure(fg_color=C["inativo"], text_color=C["inativo_txt"], font=FONTE_CORPO)
            def buscar():
                with _conectar() as conn:
                    c = conn.cursor()
                    c.execute("""
                        SELECT a.serie, a.turma, a.turno, COALESCE(SUM(r.lidos), 0) AS total
                        FROM (SELECT DISTINCT serie, turma, turno FROM alunos) a
                        LEFT JOIN ranking r ON r.serie=a.serie AND r.turma=a.turma AND r.turno=a.turno
                        GROUP BY a.serie, a.turma, a.turno ORDER BY total DESC
                    """)
                    return c.fetchall()
            def mostrar(dados):
                for w in sc.winfo_children(): w.destroy()
                if not dados:
                    ctk.CTkLabel(sc, text="Nenhuma turma cadastrada.", font=FONTE_CORPO, text_color=C["subtexto"]).pack(pady=40); return
                for i, (s, tr, tn, t) in enumerate(dados, 1):
                    card = ctk.CTkFrame(sc, fg_color=C["card"], corner_radius=10, border_width=1, border_color=C["borda"])
                    card.pack(fill="x", pady=3)
                    ctk.CTkLabel(card, text=medalhas.get(i,"—") if t>0 else "—", font=("Segoe UI", 20), width=52).pack(side="left", padx=8)
                    info = ctk.CTkFrame(card, fg_color="transparent"); info.pack(side="left", fill="x", expand=True, pady=10)
                    ctk.CTkLabel(info, text=f"Série {s or '—'}  —  Turma {tr or '—'}", font=FONTE_BOLD, text_color=C["texto"], anchor="w").pack(anchor="w", padx=4)
                    if tn: ctk.CTkLabel(info, text=f"Turno: {tn}", font=FONTE_PEQUENA, text_color=C["subtexto"], anchor="w").pack(anchor="w", padx=4)
                    cor = C["verde"] if t > 0 else C["subtexto"]
                    ctk.CTkLabel(card, text=f"{t} livro{'s' if t!=1 else ''}", font=FONTE_BOLD, text_color=cor, width=90).pack(side="right", padx=16)
            em_thread(buscar, mostrar)

        btn_alunos.configure(command=renderizar_alunos)
        btn_turmas.configure(command=renderizar_turmas)
        renderizar_alunos()

    def exportar_ranking(self):
        def _exportar():
            wb = Workbook(); ws1 = wb.active; ws1.title = "Ranking Alunos"
            ws1.append(["Posição","Aluno","Livros Lidos","Série","Turma","Turno"])
            with _conectar() as conn:
                c = conn.cursor()
                c.execute("SELECT nome, lidos, serie, turma, turno FROM ranking WHERE lidos > 0 ORDER BY lidos DESC")
                for i, row in enumerate(c.fetchall(), 1): ws1.append([i]+list(row))
                ws2 = wb.create_sheet("Ranking Turmas")
                ws2.append(["Posição","Série","Turma","Turno","Total Lidos"])
                c.execute("""
                    SELECT a.serie, a.turma, a.turno, COALESCE(SUM(r.lidos),0)
                    FROM (SELECT DISTINCT serie, turma, turno FROM alunos) a
                    LEFT JOIN ranking r ON r.serie=a.serie AND r.turma=a.turma AND r.turno=a.turno
                    GROUP BY a.serie, a.turma, a.turno ORDER BY 4 DESC
                """)
                for i, row in enumerate(c.fetchall(), 1): ws2.append([i]+list(row))
            caminho = EXPORT_DIR / f"Ranking_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            wb.save(caminho); return caminho
        em_thread(_exportar,
                  lambda p: messagebox.showinfo("Exportado", f"Ranking salvo em:\n{p}"),
                  lambda e: messagebox.showerror("Erro", str(e)))

    def zerar_ranking(self):
        if messagebox.askyesno("Atenção", "Zerar a contagem de todos os alunos?\nEssa ação não pode ser desfeita."):
            try:
                with _conectar() as conn:
                    conn.execute("UPDATE ranking SET lidos = 0"); conn.commit()
                messagebox.showinfo("Concluído", "Ranking zerado com sucesso!")
                self.tela_ranking_unificado()
            except sqlite3.Error as e:
                messagebox.showerror("Erro", str(e))

    # ── RELATÓRIOS ───────────────────────────
    def tela_relatorios(self):
        self.limpar_container()
        cabecalho_tela(self.container, "📊  Relatórios", self.mostrar_tela_principal)
        centro = ctk.CTkFrame(self.container, fg_color=C["fundo"], corner_radius=0)
        centro.pack(fill="both", expand=True)
        abas = ctk.CTkFrame(centro, fg_color="transparent"); abas.pack(fill="x", padx=30, pady=(20, 0))
        btn_c = ctk.CTkButton(abas, text="📋 Controle de Leitura", width=200, height=38, corner_radius=10,
                               fg_color=C["vermelho"], hover_color=C["vermelho_esc"], text_color="white", font=FONTE_BOLD)
        btn_c.pack(side="left", padx=(0, 8))
        btn_r = ctk.CTkButton(abas, text="🏆 Ranking Mensal", width=190, height=38, corner_radius=10,
                               fg_color=C["inativo"], hover_color="#D1D5DB", text_color=C["inativo_txt"], font=FONTE_CORPO)
        btn_r.pack(side="left", padx=(0, 8))
        btn_g = ctk.CTkButton(abas, text="🏫 Ranking Completo", width=190, height=38, corner_radius=10,
                               fg_color=C["inativo"], hover_color="#D1D5DB", text_color=C["inativo_txt"], font=FONTE_CORPO)
        btn_g.pack(side="left")
        area = ctk.CTkFrame(centro, fg_color="transparent"); area.pack(fill="both", expand=True, padx=30, pady=16)

        def ativar(b_ativo):
            for b in [btn_c, btn_r, btn_g]:
                if b == b_ativo: b.configure(fg_color=C["vermelho"], text_color="white", font=FONTE_BOLD)
                else: b.configure(fg_color=C["inativo"], text_color=C["inativo_txt"], font=FONTE_CORPO)

        def campos_relatorio(card):
            f = ctk.CTkFrame(card, fg_color="transparent"); f.pack(fill="x", padx=24, pady=(8, 0))
            ctk.CTkLabel(f, text="Série:", font=FONTE_CORPO, text_color=C["subtexto"]).pack(side="left")
            e_s = ctk.CTkEntry(f, placeholder_text="Ex: 6", **estilo_entry(80)); e_s.pack(side="left", padx=(4, 16))
            ctk.CTkLabel(f, text="Turma:", font=FONTE_CORPO, text_color=C["subtexto"]).pack(side="left")
            seg = ctk.CTkSegmentedButton(f, values=TURMAS_OPCOES, **estilo_segmented(160))
            seg.set("1"); seg.pack(side="left", padx=(4, 16))
            def atualizar(e=None):
                s = e_s.get().strip().upper()
                ops = turmas_da_serie(s) if s else TURMAS_OPCOES
                seg.configure(values=ops); seg.set(ops[0])
            e_s.bind("<FocusOut>", atualizar); e_s.bind("<KeyRelease>", atualizar)
            ctk.CTkLabel(f, text="Turno:", font=FONTE_CORPO, text_color=C["subtexto"]).pack(side="left")
            e_t = ctk.CTkEntry(f, placeholder_text="Manhã / Tarde (opcional)", **estilo_entry(180)); e_t.pack(side="left", padx=(4, 16))
            ctk.CTkLabel(f, text="Professor(a):", font=FONTE_CORPO, text_color=C["subtexto"]).pack(side="left")
            e_p = ctk.CTkEntry(f, placeholder_text="Nome", **estilo_entry(160)); e_p.pack(side="left", padx=(4, 0))
            return e_s, seg, e_t, e_p

        def aba_controle():
            for w in area.winfo_children(): w.destroy(); ativar(btn_c)
            card = ctk.CTkFrame(area, fg_color=C["card"], corner_radius=14, border_width=1, border_color=C["borda"]); card.pack(fill="x")
            ctk.CTkLabel(card, text="Gerar planilha de controle de leitura por turma", font=FONTE_SECAO, text_color=C["vermelho"]).pack(anchor="w", padx=24, pady=(20, 4))
            ctk.CTkLabel(card, text="Lista todos os alunos da turma com colunas semanais e total de livros lidos.", font=FONTE_PEQUENA, text_color=C["subtexto"], wraplength=650).pack(anchor="w", padx=24, pady=(0, 12))
            separador(card, pady=(0, 12))
            e_s, seg, e_t, e_p = campos_relatorio(card)
            lbl = ctk.CTkLabel(card, text="", font=FONTE_PEQUENA, text_color=C["verde"], wraplength=650); lbl.pack(anchor="w", padx=24, pady=(10, 0))
            ctk.CTkButton(card, text="📥 Gerar Planilha Excel", **estilo_btn_verde(220, 42),
                          command=lambda: self._gerar_controle(e_s, seg, e_t, e_p, lbl)).pack(anchor="w", padx=24, pady=(12, 24))

        def aba_ranking_mensal():
            for w in area.winfo_children(): w.destroy(); ativar(btn_r)
            card = ctk.CTkFrame(area, fg_color=C["card"], corner_radius=14, border_width=1, border_color=C["borda"]); card.pack(fill="x")
            ctk.CTkLabel(card, text="Gerar planilha de ranking mensal por turma", font=FONTE_SECAO, text_color=C["vermelho"]).pack(anchor="w", padx=24, pady=(20, 4))
            ctk.CTkLabel(card, text="Formato da escola: aluno na vertical, semanas na horizontal, X em azul e total destacado.", font=FONTE_PEQUENA, text_color=C["subtexto"], wraplength=650).pack(anchor="w", padx=24, pady=(0, 12))
            separador(card, pady=(0, 12))
            f = ctk.CTkFrame(card, fg_color="transparent"); f.pack(fill="x", padx=24, pady=(8, 0))
            ctk.CTkLabel(f, text="Série:", font=FONTE_CORPO, text_color=C["subtexto"]).pack(side="left")
            e_s2 = ctk.CTkEntry(f, placeholder_text="Ex: 5", **estilo_entry(80)); e_s2.pack(side="left", padx=(4, 16))
            ctk.CTkLabel(f, text="Turma:", font=FONTE_CORPO, text_color=C["subtexto"]).pack(side="left")
            seg2 = ctk.CTkSegmentedButton(f, values=TURMAS_OPCOES, **estilo_segmented(160))
            seg2.set("1"); seg2.pack(side="left", padx=(4, 16))
            def atualizar2(e=None):
                s = e_s2.get().strip().upper()
                ops = turmas_da_serie(s) if s else TURMAS_OPCOES
                seg2.configure(values=ops); seg2.set(ops[0])
            e_s2.bind("<FocusOut>", atualizar2); e_s2.bind("<KeyRelease>", atualizar2)
            ctk.CTkLabel(f, text="Mês:", font=FONTE_CORPO, text_color=C["subtexto"]).pack(side="left")
            seg_mes = ctk.CTkSegmentedButton(f, values=list(MESES_PT.values()),
                                              selected_color=C["verde"], selected_hover_color=C["verde_esc"],
                                              unselected_color=C["inativo"], unselected_hover_color="#D1D5DB",
                                              text_color="white", font=("Segoe UI", 11, "bold"),
                                              corner_radius=8, width=600, height=34)
            seg_mes.set(MESES_PT[datetime.now().month]); seg_mes.pack(side="left", padx=(4, 16))
            ctk.CTkLabel(f, text="Professor(a):", font=FONTE_CORPO, text_color=C["subtexto"]).pack(side="left")
            e_p2 = ctk.CTkEntry(f, placeholder_text="Nome", **estilo_entry(160)); e_p2.pack(side="left", padx=(4, 0))
            lbl2 = ctk.CTkLabel(card, text="", font=FONTE_PEQUENA, text_color=C["verde"], wraplength=650); lbl2.pack(anchor="w", padx=24, pady=(10, 0))
            ctk.CTkButton(card, text="📥 Gerar Ranking Mensal", **estilo_btn_verde(230, 42),
                          command=lambda: self._gerar_ranking_mensal(e_s2, seg2, seg_mes, e_p2, lbl2)).pack(anchor="w", padx=24, pady=(12, 24))

        def aba_geral():
            for w in area.winfo_children(): w.destroy(); ativar(btn_g)
            card = ctk.CTkFrame(area, fg_color=C["card"], corner_radius=14, border_width=1, border_color=C["borda"]); card.pack(fill="x")
            ctk.CTkLabel(card, text="Exportar ranking completo de todas as turmas", font=FONTE_SECAO, text_color=C["vermelho"]).pack(anchor="w", padx=24, pady=(20, 4))
            ctk.CTkLabel(card, text="Planilha com Ranking Alunos e Ranking Turmas. Todas as turmas aparecem, mesmo com 0 livros.", font=FONTE_PEQUENA, text_color=C["subtexto"], wraplength=650).pack(anchor="w", padx=24, pady=(0, 12))
            separador(card, pady=(0, 12))
            ctk.CTkButton(card, text="📥 Exportar Ranking Completo", **estilo_btn_verde(240, 42),
                          command=self.exportar_ranking).pack(anchor="w", padx=24, pady=(12, 24))

        btn_c.configure(command=aba_controle)
        btn_r.configure(command=aba_ranking_mensal)
        btn_g.configure(command=aba_geral)
        aba_controle()

    def _gerar_controle(self, e_serie, seg_turma, e_turno, e_prof, lbl_status):
        serie = e_serie.get().strip().upper(); turma = seg_turma.get().strip()
        turno = e_turno.get().strip(); prof = e_prof.get().strip()
        if not serie: messagebox.showwarning("Atenção", "Série é obrigatória!"); return
        def _gerar():
            with _conectar() as conn:
                c = conn.cursor()
                if turno: c.execute("SELECT nome FROM alunos WHERE serie=? AND turma=? AND turno=? ORDER BY nome", (serie, turma, turno.upper()))
                else: c.execute("SELECT nome FROM alunos WHERE serie=? AND turma=? ORDER BY nome", (serie, turma))
                alunos = [r[0] for r in c.fetchall()]
                if not alunos: return None, serie, turma
                if turno: c.execute("SELECT nome, lidos FROM ranking WHERE serie=? AND turma=? AND turno=?", (serie, turma, turno.upper()))
                else: c.execute("SELECT nome, lidos FROM ranking WHERE serie=? AND turma=?", (serie, turma))
                dict_lidos = {r[0]: r[1] for r in c.fetchall()}
            return alunos, dict_lidos
        def _salvar(res):
            if res[0] is None: messagebox.showinfo("Info", f"Nenhum aluno encontrado na {res[1]}ª série, turma {res[2]}."); return
            alunos, dict_lidos = res
            turno_label = turno.upper() if turno else "TODOS OS TURNOS"
            wb = Workbook(); ws = wb.active; ws.title = f"{serie}-T{turma}"
            ws.merge_cells("A1:F1"); ws["A1"] = f"CONTROLE DE LEITURA — {serie}ª SÉRIE  |  TURMA {turma}  |  {turno_label}   —   {MESES_PT[datetime.now().month]}/{datetime.now().year}"
            ws["A1"].font = Font(bold=True, size=13); ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A1"].fill = PatternFill("solid", fgColor="FFCC00"); ws.row_dimensions[1].height = 28; ws.row_dimensions[2].height = 6
            for col, h in enumerate(["ALUNO","1ª Semana","2ª Semana","3ª Semana","4ª Semana","TOTAL LIDOS"], 1):
                cell = ws.cell(3, col, h); cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill("solid", fgColor="D9D9D9")
            ws.row_dimensions[3].height = 20
            vf = PatternFill("solid", fgColor="E8F5E9"); thin = thin_border()
            for i, aluno in enumerate(alunos, 4):
                lidos = dict_lidos.get(aluno, 0); ws.cell(i, 1, aluno).font = Font(size=11)
                ct = ws.cell(i, 6, lidos); ct.font = Font(bold=True, size=11); ct.alignment = Alignment(horizontal="center")
                for col in range(2, 6):
                    c2 = ws.cell(i, col, "X" if lidos > 0 else ""); c2.alignment = Alignment(horizontal="center")
                    if lidos > 0: c2.fill = vf
                ws.row_dimensions[i].height = 18
            for row in ws.iter_rows(min_row=3, max_row=3+len(alunos), min_col=1, max_col=6):
                for cell in row: cell.border = thin
            ws.column_dimensions["A"].width = 36
            for col in ["B","C","D","E"]: ws.column_dimensions[col].width = 14
            ws.column_dimensions["F"].width = 16
            caminho = EXPORT_DIR / f"Controle_{serie}-{turma}_{turno}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            try: wb.save(caminho); lbl_status.configure(text=f"✅  Arquivo gerado!\n{caminho}")
            except Exception as e: messagebox.showerror("Erro ao salvar", str(e))
        em_thread(_gerar, _salvar, lambda e: messagebox.showerror("Erro", str(e)))

    def _gerar_ranking_mensal(self, e_serie, seg_turma, seg_mes, e_prof, lbl_status):
        serie = e_serie.get().strip().upper(); turma = seg_turma.get().strip()
        mes_nome = seg_mes.get().strip(); prof = e_prof.get().strip()
        if not serie: messagebox.showwarning("Atenção", "Série é obrigatória!"); return
        def _gerar():
            with _conectar() as conn:
                c = conn.cursor()
                c.execute("SELECT nome FROM alunos WHERE serie=? AND turma=? ORDER BY nome", (serie, turma))
                alunos = [r[0] for r in c.fetchall()]
                if not alunos: return None, serie, turma
                c.execute("SELECT nome, lidos FROM ranking WHERE serie=? AND turma=?", (serie, turma))
                dict_lidos = {r[0]: r[1] for r in c.fetchall()}
            return alunos, dict_lidos
        def _salvar(res):
            if res[0] is None: messagebox.showinfo("Info", f"Nenhum aluno encontrado na {res[1]}ª série, turma {res[2]}."); return
            alunos, dict_lidos = res
            wb = Workbook(); ws = wb.active; ws.title = f"Ranking {mes_nome[:3]}"
            ws.merge_cells("A1:K1"); ws["A1"].value = "CENTRO DE EXC. MUL. D. JOÃO J DA MOTA E ALBUQUERQUE"
            ws["A1"].font = Font(bold=True, size=11); ws["A1"].alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 20
            ws["A2"] = f"Mês: {mes_nome}"; ws.merge_cells("A2:B2"); ws["A2"].font = Font(bold=True, size=11)
            ws["C2"] = f"TURMA: {serie}º ANO {turma}"; ws.merge_cells("C2:F2"); ws["C2"].font = Font(bold=True, size=11); ws["C2"].alignment = Alignment(horizontal="center")
            ws["G2"] = f"PROFESSORA: {prof}" if prof else "PROFESSORA:"; ws.merge_cells("G2:K2"); ws["G2"].font = Font(bold=True, size=11); ws.row_dimensions[2].height = 20
            ws.merge_cells("C3:F3"); ws["C3"] = f"Livros lidos {datetime.now().year}"; ws["C3"].font = Font(bold=True, size=10); ws["C3"].alignment = Alignment(horizontal="center"); ws.row_dimensions[3].height = 16
            cab_fill = PatternFill("solid", fgColor="D9D9D9"); amarelo_f = PatternFill("solid", fgColor="FFFF99")
            for col_i, h in enumerate(["ALUNA(O)","1ª\nSEMANA","2ª\nSEMANA","3ª\nSEMANA","4ª\nSEMANA","TOTAL"], 1):
                cell = ws.cell(4, col_i, h); cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = cab_fill; cell.border = thin_border()
            ws.row_dimensions[4].height = 28
            for ci, w in enumerate([32, 9, 9, 9, 9, 8], 1): ws.column_dimensions[get_column_letter(ci)].width = w
            rosa = PatternFill("solid", fgColor="FFE4E1"); branco = PatternFill("solid", fgColor="FFFFFF")
            for row_i, aluno in enumerate(alunos, 5):
                lidos = dict_lidos.get(aluno, 0); fill = rosa if (row_i % 2 == 0) else branco
                ca = ws.cell(row_i, 1, aluno); ca.font = Font(size=10); ca.alignment = Alignment(vertical="center"); ca.fill = fill; ca.border = thin_border()
                for col_s in range(2, 6):
                    cs = ws.cell(row_i, col_s, "X" if lidos > 0 else ""); cs.font = Font(bold=True, size=11, color="0000CD")
                    cs.alignment = Alignment(horizontal="center", vertical="center"); cs.fill = fill; cs.border = thin_border()
                ct = ws.cell(row_i, 6, lidos if lidos > 0 else ""); ct.font = Font(bold=True, size=11)
                ct.alignment = Alignment(horizontal="center", vertical="center"); ct.fill = amarelo_f if lidos > 0 else fill; ct.border = thin_border()
                ws.row_dimensions[row_i].height = 16
            caminho = EXPORT_DIR / f"RankingMensal_{serie}-{turma}_{mes_nome}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            try: wb.save(caminho); lbl_status.configure(text=f"✅  Ranking mensal gerado!\n{caminho}")
            except Exception as e: messagebox.showerror("Erro ao salvar", str(e))
        em_thread(_gerar, _salvar, lambda e: messagebox.showerror("Erro", str(e)))

    def exportar_dados(self, tipo):
        def _exportar():
            wb = Workbook(); ws = wb.active
            with _conectar() as conn:
                c = conn.cursor()
                if tipo == "acervo":
                    ws.title = "Acervo"
                    colunas_db = _colunas_acervo()
                    headers = ["Título","Autor","Gênero"]
                    cols_sql = ["titulo","autor","genero"]
                    if "edição"  in colunas_db: headers.append("Edição");  cols_sql.append('"edição"')
                    if "codigo"  in colunas_db: headers.append("Código");  cols_sql.append("codigo")
                    if "estante" in colunas_db: headers.append("Estante"); cols_sql.append("estante")
                    if "fileira" in colunas_db: headers.append("Fileira"); cols_sql.append("fileira")
                    ws.append(headers)
                    c.execute(f"SELECT {', '.join(cols_sql)} FROM acervo ORDER BY titulo")
                else:
                    ws.title = "Alunos"; ws.append(["Nome","Série","Turma","Turno"])
                    c.execute("SELECT nome, serie, turma, turno FROM alunos ORDER BY nome")
                for row in c.fetchall(): ws.append(list(row))
            caminho = EXPORT_DIR / f"Export_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            wb.save(caminho); return caminho
        em_thread(_exportar,
                  lambda p: messagebox.showinfo("Sucesso", f"{tipo.capitalize()} exportado!\n\n{p}"),
                  lambda e: messagebox.showerror("Erro", str(e)))


if __name__ == "__main__":
    app = BibliotecaApp()
    app.mainloop()